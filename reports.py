from datetime import datetime, timedelta
from decimal import Decimal
from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from flask_login import login_required
from sqlalchemy import func
import io
import os
from io import BytesIO
import json

# استيراد المكتبات للتصدير
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from models import db, Invoice, InvoiceItem, Product, TaxReport, TaxType, SystemSettings
from forms import ReportForm
from auth import permission_required

# استيراد مولد PDF
try:
    from pdf_generator import pdf_generator
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/reports')
@login_required
@permission_required('view_reports')
def reports_dashboard():
    """لوحة تحكم التقارير"""
    # الحصول على الفلاتر من الطلب
    period = request.args.get('period', 'current_month')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # تحديد الفترة الزمنية
    now = datetime.utcnow()
    if period == 'current_month':
        filter_start = now.replace(day=1)
        filter_end = now
    elif period == 'last_month':
        last_month = now.replace(day=1) - timedelta(days=1)
        filter_start = last_month.replace(day=1)
        filter_end = now.replace(day=1) - timedelta(days=1)
    elif period == 'current_quarter':
        quarter_start_month = ((now.month - 1) // 3) * 3 + 1
        filter_start = now.replace(month=quarter_start_month, day=1)
        filter_end = now
    elif period == 'current_year':
        filter_start = now.replace(month=1, day=1)
        filter_end = now
    elif period == 'custom' and start_date and end_date:
        filter_start = datetime.strptime(start_date, '%Y-%m-%d')
        filter_end = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        # افتراضي: الشهر الحالي
        filter_start = now.replace(day=1)
        filter_end = now
    
    # استعلام الفواتير حسب الفترة المحددة
    filtered_invoices = Invoice.query.filter(
        Invoice.invoice_date >= filter_start,
        Invoice.invoice_date <= filter_end,
        Invoice.is_cancelled == False
    ).all()
    
    # إحصائيات سريعة
    total_invoices = len(filtered_invoices)
    total_products = Product.query.count()
    
    # حساب إجماليات الضرائب والمبيعات للفترة المحددة
    total_vat = sum(invoice.vat_amount for invoice in filtered_invoices)
    total_withholding = sum(invoice.withholding_amount for invoice in filtered_invoices)
    total_sales = sum(invoice.subtotal for invoice in filtered_invoices)
    
    # إحصائيات الشهر الحالي
    current_month = datetime.utcnow().replace(day=1)
    monthly_invoices = Invoice.query.filter(
        Invoice.invoice_date >= current_month,
        Invoice.is_cancelled == False
    ).all()
    
    monthly_revenue = sum(invoice.total_amount for invoice in monthly_invoices)
    monthly_vat = sum(invoice.vat_amount for invoice in monthly_invoices)
    monthly_withholding = sum(invoice.withholding_amount for invoice in monthly_invoices)
    
    # أفضل المنتجات مبيعاً
    top_products_query = db.session.query(
        Product.name,
        func.sum(InvoiceItem.quantity).label('total_quantity'),
        func.sum(InvoiceItem.quantity * InvoiceItem.unit_price).label('total_revenue')
    ).join(InvoiceItem).join(Invoice).filter(
        Invoice.is_cancelled == False
    ).group_by(Product.id).order_by(
        func.sum(InvoiceItem.quantity * InvoiceItem.unit_price).desc()
    ).limit(5).all()
    
    top_products = [
        {
            'name': product[0],
            'quantity': product[1],
            'revenue': product[2]
        }
        for product in top_products_query
    ]
    
    # إحصائيات السنة الحالية
    current_year = datetime.utcnow().replace(month=1, day=1)
    yearly_invoices = Invoice.query.filter(
        Invoice.invoice_date >= current_year,
        Invoice.is_cancelled == False
    ).all()
    
    yearly_revenue = sum(invoice.total_amount for invoice in yearly_invoices)
    yearly_vat = sum(invoice.vat_amount for invoice in yearly_invoices)
    yearly_withholding = sum(invoice.withholding_amount for invoice in yearly_invoices)
    
    # بيانات للرسوم البيانية (آخر 12 شهر)
    monthly_labels = []
    monthly_revenues = []
    monthly_vat_amounts = []
    monthly_withholding_amounts = []
    
    for i in range(12):
        month_start = (datetime.utcnow().replace(day=1) - timedelta(days=32*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        month_invoices = Invoice.query.filter(
            Invoice.invoice_date >= month_start,
            Invoice.invoice_date <= month_end,
            Invoice.is_cancelled == False
        ).all()
        
        monthly_labels.insert(0, month_start.strftime('%Y/%m'))
        monthly_revenues.insert(0, sum(inv.total_amount for inv in month_invoices))
        monthly_vat_amounts.insert(0, sum(inv.vat_amount for inv in month_invoices))
        monthly_withholding_amounts.insert(0, sum(inv.withholding_amount for inv in month_invoices))
    
    # إحصائيات إضافية
    month_stats = {
        'revenue': monthly_revenue,
        'vat': monthly_vat,
        'withholding': monthly_withholding,
        'invoices_count': len(monthly_invoices)
    }
    
    year_stats = {
        'revenue': yearly_revenue,
        'vat': yearly_vat,
        'withholding': yearly_withholding,
        'invoices_count': len(yearly_invoices)
    }
    
    return render_template('reports/dashboard.html',
                         total_invoices=total_invoices,
                         total_products=total_products,
                         total_vat=total_vat,
                         total_withholding=total_withholding,
                         total_sales=total_sales,
                         monthly_labels=monthly_labels,
                         monthly_vat=monthly_vat_amounts,
                         monthly_withholding=monthly_withholding_amounts,
                         monthly_revenues=monthly_revenues,
                         month_stats=month_stats,
                         year_stats=year_stats,
                         top_products=top_products,
                         period=period,
                         start_date=start_date,
                         end_date=end_date,
                         period_text=get_period_text(period, start_date, end_date))

def get_period_text(period, start_date, end_date):
    """الحصول على نص الفترة الزمنية"""
    if period == 'current_month':
        return 'الشهر الحالي'
    elif period == 'last_month':
        return 'الشهر الماضي'
    elif period == 'current_quarter':
        return 'الربع الحالي'
    elif period == 'current_year':
        return 'السنة الحالية'
    elif period == 'custom' and start_date and end_date:
        return f'من {start_date} إلى {end_date}'
    else:
        return 'الشهر الحالي'

@reports_bp.route('/reports/generate', methods=['GET', 'POST'])
@login_required
@permission_required('view_reports')
def generate_report():
    """إنشاء تقرير مخصص"""
    form = ReportForm()
    
    if form.validate_on_submit():
        # تحديد الفترة الزمنية
        period_start = form.period_start.data
        period_end = form.period_end.data
        
        # استعلام الفواتير
        query = Invoice.query.filter(
            Invoice.invoice_date >= period_start,
            Invoice.invoice_date <= period_end
        )
        
        if not form.include_cancelled.data:
            query = query.filter(Invoice.is_cancelled == False)
        
        invoices = query.all()
        
        # حساب الإجماليات
        report_data = calculate_report_totals(invoices)
        report_data.update({
            'period_start': period_start,
            'period_end': period_end,
            'report_type': form.report_type.data,
            'include_cancelled': form.include_cancelled.data,
            'generated_by': current_user.username,
            'generated_at': datetime.now(),
            'invoices': invoices
        })
        
        # حفظ التقرير في قاعدة البيانات
        tax_report = TaxReport(
            report_type=form.report_type.data,
            period_start=period_start,
            period_end=period_end,
            total_sales=report_data['total_sales'],
            total_vat=report_data['total_vat'],
            total_withholding=report_data['total_withholding'],
            generated_by=current_user.id
        )
        db.session.add(tax_report)
        db.session.commit()
        
        # تصدير التقرير
        if form.export_format.data in ['pdf', 'both']:
            if REPORTLAB_AVAILABLE:
                pdf_file = export_report_pdf(report_data, tax_report.id)
                tax_report.file_path = pdf_file
                db.session.commit()
            else:
                flash('مكتبة PDF غير متوفرة. يرجى تثبيت reportlab.', 'warning')
        
        if form.export_format.data in ['excel', 'both']:
            if OPENPYXL_AVAILABLE:
                excel_file = export_report_excel(report_data, tax_report.id)
                if not tax_report.file_path:
                    tax_report.file_path = excel_file
                db.session.commit()
            else:
                flash('مكتبة Excel غير متوفرة. يرجى تثبيت openpyxl.', 'warning')
        
        flash('تم إنشاء التقرير بنجاح.', 'success')
        return redirect(url_for('reports.view_report', report_id=tax_report.id))
    
    return render_template('reports/generate.html', form=form)

@reports_bp.route('/reports/<int:report_id>')
@login_required
@permission_required('view_reports')
def view_report(report_id):
    """عرض تقرير محفوظ"""
    report = TaxReport.query.get_or_404(report_id)
    
    # إعادة حساب البيانات للعرض
    invoices = Invoice.query.filter(
        Invoice.invoice_date >= report.period_start,
        Invoice.invoice_date <= report.period_end,
        Invoice.is_cancelled == False
    ).all()
    
    report_data = calculate_report_totals(invoices)
    report_data.update({
        'report': report,
        'invoices': invoices
    })
    
    return render_template('reports/view.html', **report_data)

@reports_bp.route('/reports/list')
@login_required
@permission_required('view_reports')
def reports_list():
    """قائمة التقارير المحفوظة"""
    page = request.args.get('page', 1, type=int)
    reports = TaxReport.query.order_by(TaxReport.generated_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    return render_template('reports/list.html', reports=reports)

@reports_bp.route('/reports/<int:report_id>/download/<format>')
@login_required
@permission_required('export_reports')
def download_report(report_id, format):
    """تحميل التقرير"""
    report = TaxReport.query.get_or_404(report_id)
    
    # إعادة إنشاء البيانات
    invoices = Invoice.query.filter(
        Invoice.invoice_date >= report.period_start,
        Invoice.invoice_date <= report.period_end,
        Invoice.is_cancelled == False
    ).all()
    
    report_data = calculate_report_totals(invoices)
    report_data.update({
        'period_start': report.period_start,
        'period_end': report.period_end,
        'report_type': report.report_type,
        'generated_by': report.generated_by_user.username,
        'generated_at': report.generated_at,
        'invoices': invoices
    })
    
    if format == 'pdf' and REPORTLAB_AVAILABLE:
        pdf_buffer = create_pdf_report(report_data)
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f'tax_report_{report_id}.pdf',
            mimetype='application/pdf'
        )
    elif format == 'excel' and OPENPYXL_AVAILABLE:
        excel_buffer = create_excel_report(report_data)
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f'tax_report_{report_id}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        flash('تنسيق التصدير غير مدعوم.', 'error')
        return redirect(url_for('reports.view_report', report_id=report_id))

@reports_bp.route('/reports/vat')
@login_required
@permission_required('view_reports')
def vat_report():
    """تقرير ضريبة القيمة المضافة"""
    # استعلام الفواتير التي تحتوي على ضريبة قيمة مضافة
    invoices = Invoice.query.filter(
        Invoice.vat_amount > 0,
        Invoice.is_cancelled == False
    ).order_by(Invoice.invoice_date.desc()).all()
    
    # حساب الإجماليات - المبلغ الخاضع للضريبة فقط
    total_taxable_sales = 0
    total_vat_amount = 0
    
    for invoice in invoices:
        if invoice.vat_amount > 0:
            # حساب المبلغ الخاضع لضريبة القيمة المضافة
            vat_base = float(invoice.vat_amount / Decimal('0.14'))
            total_taxable_sales += vat_base
            total_vat_amount += float(invoice.vat_amount)
    
    # الحصول على بيانات الشركة من الإعدادات
    company_name = SystemSettings.get_setting('company_name', 'اسم الشركة')
    tax_number = SystemSettings.get_setting('tax_number', '000000000')
    company_address = SystemSettings.get_setting('company_address', 'عنوان الشركة')
    
    # تصدير إذا طُلب
    export_format = request.args.get('export')
    if export_format == 'pdf' and PDF_AVAILABLE:
        try:
            # تحضير البيانات للـ PDF
            pdf_data = {
                'company_name': company_name,
                'tax_number': tax_number,
                'company_address': company_address,
                'invoices': invoices,
                'total_taxable_sales': total_taxable_sales,
                'total_vat_amount': total_vat_amount
            }
            
            # إنشاء PDF
            pdf_buffer = pdf_generator.generate_vat_report_pdf(pdf_data)
            
            # إرسال الملف
            filename = f"vat_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        except Exception as e:
            flash(f'خطأ في إنشاء PDF: {str(e)}', 'error')
    elif export_format == 'pdf':
        flash('تصدير PDF غير متاح حالياً', 'warning')
    elif export_format == 'excel':
        flash('تصدير Excel قيد التطوير', 'info')
    
    return render_template('reports/vat_report.html',
                         invoices=invoices,
                         total_taxable_sales=total_taxable_sales,
                         total_vat_amount=total_vat_amount,
                         period_text='جميع الفترات',
                         current_date=datetime.utcnow(),
                         company_name=company_name,
                         tax_number=tax_number,
                         company_address=company_address)

@reports_bp.route('/reports/withholding')
@login_required
@permission_required('view_reports')
def withholding_report():
    """تقرير ضريبة الخصم والإضافة"""
    # استعلام الفواتير التي تحتوي على ضريبة خصم وإضافة
    invoices = Invoice.query.filter(
        Invoice.withholding_amount > 0,
        Invoice.is_cancelled == False
    ).order_by(Invoice.invoice_date.desc()).all()
    
    # حساب الإجماليات - المبلغ الخاضع للضريبة فقط
    total_taxable_sales = 0
    total_withholding_amount = 0
    
    for invoice in invoices:
        if invoice.withholding_amount > 0:
            # حساب المبلغ الخاضع لضريبة الخصم والإضافة
            withholding_base = float(invoice.withholding_amount / Decimal('0.05'))
            total_taxable_sales += withholding_base
            total_withholding_amount += float(invoice.withholding_amount)
    
    # الحصول على بيانات الشركة من الإعدادات
    company_name = SystemSettings.get_setting('company_name', 'اسم الشركة')
    tax_number = SystemSettings.get_setting('tax_number', '000000000')
    company_address = SystemSettings.get_setting('company_address', 'عنوان الشركة')
    
    # تصدير إذا طُلب
    export_format = request.args.get('export')
    if export_format == 'pdf' and PDF_AVAILABLE:
        try:
            # تحضير البيانات للـ PDF
            pdf_data = {
                'company_name': company_name,
                'tax_number': tax_number,
                'company_address': company_address,
                'invoices': invoices,
                'total_taxable_sales': total_taxable_sales,
                'total_withholding_amount': total_withholding_amount
            }
            
            # إنشاء PDF
            pdf_buffer = pdf_generator.generate_withholding_report_pdf(pdf_data)
            
            # إرسال الملف
            filename = f"withholding_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        except Exception as e:
            flash(f'خطأ في إنشاء PDF: {str(e)}', 'error')
    elif export_format == 'pdf':
        flash('تصدير PDF غير متاح حالياً', 'warning')
    elif export_format == 'excel':
        flash('تصدير Excel قيد التطوير', 'info')
    
    return render_template('reports/withholding_report.html',
                         invoices=invoices,
                         total_taxable_sales=total_taxable_sales,
                         total_withholding_amount=total_withholding_amount,
                         period_text='جميع الفترات',
                         current_date=datetime.utcnow(),
                         company_name=company_name,
                         tax_number=tax_number,
                         company_address=company_address)

@reports_bp.route('/reports/sales')
@login_required
@permission_required('view_reports')
def sales_report():
    """تقرير المبيعات الصافية"""
    # استعلام جميع الفواتير النشطة
    invoices = Invoice.query.filter(
        Invoice.is_cancelled == False
    ).order_by(Invoice.invoice_date.desc()).all()
    
    # حساب الإجماليات
    total_sales = sum(invoice.subtotal for invoice in invoices)
    total_vat = sum(invoice.vat_amount for invoice in invoices)
    total_withholding = sum(invoice.withholding_amount for invoice in invoices)
    total_taxes = total_vat + total_withholding
    
    # الحصول على بيانات الشركة من الإعدادات
    company_name = SystemSettings.get_setting('company_name', 'اسم الشركة')
    tax_number = SystemSettings.get_setting('tax_number', '000000000')
    company_address = SystemSettings.get_setting('company_address', 'عنوان الشركة')
    
    # تصدير إذا طُلب
    export_format = request.args.get('export')
    if export_format == 'pdf' and PDF_AVAILABLE:
        try:
            # تحضير البيانات للـ PDF
            pdf_data = {
                'company_name': company_name,
                'tax_number': tax_number,
                'company_address': company_address,
                'invoices': invoices,
                'total_sales': total_sales,
                'total_vat': total_vat,
                'total_withholding': total_withholding,
                'total_taxes': total_taxes
            }
            
            # إنشاء PDF
            pdf_buffer = pdf_generator.generate_sales_report_pdf(pdf_data)
            
            # إرسال الملف
            filename = f"sales_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        except Exception as e:
            flash(f'خطأ في إنشاء PDF: {str(e)}', 'error')
    elif export_format == 'pdf':
        flash('تصدير PDF غير متاح حالياً', 'warning')
    elif export_format == 'excel':
        flash('تصدير Excel قيد التطوير', 'info')
    
    return render_template('reports/sales_report.html',
                         invoices=invoices,
                         total_sales=total_sales,
                         total_vat=total_vat,
                         total_withholding=total_withholding,
                         total_taxes=total_taxes,
                         period_text='جميع الفترات',
                         current_date=datetime.utcnow(),
                         company_name=company_name,
                         tax_number=tax_number,
                         company_address=company_address)

@reports_bp.route('/reports/comprehensive')
@login_required
@permission_required('view_reports')
def comprehensive_report():
    """التقرير الشامل"""
    # استعلام جميع الفواتير النشطة
    invoices = Invoice.query.filter(
        Invoice.is_cancelled == False
    ).order_by(Invoice.invoice_date.desc()).all()
    
    # حساب الإجماليات
    total_invoices = len(invoices)
    total_sales = sum(invoice.subtotal for invoice in invoices)
    total_vat = sum(invoice.vat_amount for invoice in invoices)
    total_withholding = sum(invoice.withholding_amount for invoice in invoices)
    total_taxes = total_vat + total_withholding
    
    # إحصائيات تفصيلية
    vat_invoices = [inv for inv in invoices if inv.vat_amount > 0]
    withholding_invoices = [inv for inv in invoices if inv.withholding_amount > 0]
    
    vat_invoices_count = len(vat_invoices)
    withholding_invoices_count = len(withholding_invoices)
    
    # حساب المبيعات الخاضعة للضريبة بشكل صحيح
    vat_taxable_sales = 0
    withholding_taxable_sales = 0
    
    for inv in vat_invoices:
        if inv.vat_amount > 0:
            vat_taxable_sales += float(inv.vat_amount / Decimal('0.14'))
    
    for inv in withholding_invoices:
        if inv.withholding_amount > 0:
            withholding_taxable_sales += float(inv.withholding_amount / Decimal('0.05'))
    
    # الحصول على بيانات الشركة من الإعدادات
    company_name = SystemSettings.get_setting('company_name', 'اسم الشركة')
    tax_number = SystemSettings.get_setting('tax_number', '000000000')
    company_address = SystemSettings.get_setting('company_address', 'عنوان الشركة')
    
    # تصدير إذا طُلب
    export_format = request.args.get('export')
    if export_format == 'pdf' and PDF_AVAILABLE:
        try:
            # تحضير البيانات للـ PDF
            pdf_data = {
                'company_name': company_name,
                'tax_number': tax_number,
                'company_address': company_address,
                'invoices': invoices,
                'total_invoices': total_invoices,
                'total_sales': total_sales,
                'total_vat': total_vat,
                'total_withholding': total_withholding,
                'total_taxes': total_taxes,
                'vat_invoices_count': vat_invoices_count,
                'withholding_invoices_count': withholding_invoices_count,
                'vat_taxable_sales': vat_taxable_sales,
                'withholding_taxable_sales': withholding_taxable_sales
            }
            
            # إنشاء PDF
            pdf_buffer = pdf_generator.generate_comprehensive_report_pdf(pdf_data)
            
            # إرسال الملف
            filename = f"comprehensive_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        except Exception as e:
            flash(f'خطأ في إنشاء PDF: {str(e)}', 'error')
    elif export_format == 'pdf':
        flash('تصدير PDF غير متاح حالياً', 'warning')
    elif export_format == 'excel':
        flash('تصدير Excel قيد التطوير', 'info')
    
    return render_template('reports/comprehensive_report.html',
                         invoices=invoices,
                         total_invoices=total_invoices,
                         total_sales=total_sales,
                         total_vat=total_vat,
                         total_withholding=total_withholding,
                         total_taxes=total_taxes,
                         vat_invoices_count=vat_invoices_count,
                         withholding_invoices_count=withholding_invoices_count,
                         vat_taxable_sales=vat_taxable_sales,
                         withholding_taxable_sales=withholding_taxable_sales,
                         period_text='جميع الفترات',
                         current_date=datetime.utcnow(),
                         company_name=company_name,
                         tax_number=tax_number,
                         company_address=company_address)

@reports_bp.route('/reports/tax-declaration')
@login_required
@permission_required('view_reports')
def tax_declaration():
    """إنشاء إقرار ضريبي"""
    # الحصول على جميع الفواتير النشطة
    invoices = Invoice.query.filter_by(is_cancelled=False).all()
    
    # حساب الإجماليات للإقرار بشكل محاسبي صحيح
    total_sales = sum(inv.subtotal for inv in invoices)  # إجمالي المبيعات
    
    # حساب المبيعات الخاضعة لضريبة القيمة المضافة فقط
    # (المبيعات التي عليها ضريبة 14% مقسومة على 1.14 للحصول على القيمة قبل الضريبة)
    total_vat_sales = 0
    total_vat_amount = 0
    
    # حساب المبيعات الخاضعة لضريبة الخصم والإضافة فقط  
    # (المبيعات التي عليها ضريبة 5% مقسومة على 1.05 للحصول على القيمة قبل الضريبة)
    total_withholding_sales = 0
    total_withholding_amount = 0
    
    for invoice in invoices:
        if invoice.vat_amount > 0:
            # حساب المبيعات الخاضعة لضريبة القيمة المضافة
            vat_base = invoice.vat_amount / Decimal('0.14')  # القيمة قبل الضريبة
            total_vat_sales += float(vat_base)
            total_vat_amount += float(invoice.vat_amount)
            
        if invoice.withholding_amount > 0:
            # حساب المبيعات الخاضعة لضريبة الخصم والإضافة
            withholding_base = invoice.withholding_amount / Decimal('0.05')  # القيمة قبل الضريبة
            total_withholding_sales += float(withholding_base)
            total_withholding_amount += float(invoice.withholding_amount)
    
    # إحصائيات شهرية للسنة الحالية
    current_year = datetime.utcnow().year
    monthly_data = []
    
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1)
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = datetime(current_year, month + 1, 1) - timedelta(days=1)
        
        month_invoices = Invoice.query.filter(
            Invoice.invoice_date >= month_start,
            Invoice.invoice_date <= month_end,
            Invoice.is_cancelled == False
        ).all()
        
        # حساب المبيعات الشهرية بشكل محاسبي صحيح
        month_total_sales = sum(inv.subtotal for inv in month_invoices)
        
        # حساب المبيعات الخاضعة للضرائب لكل شهر
        month_vat_sales = 0
        month_vat_amount = 0
        month_withholding_sales = 0
        month_withholding_amount = 0
        
        for invoice in month_invoices:
            if invoice.vat_amount > 0:
                month_vat_sales += float(invoice.vat_amount / Decimal('0.14'))  # القيمة قبل الضريبة
                month_vat_amount += float(invoice.vat_amount)
                
            if invoice.withholding_amount > 0:
                month_withholding_sales += float(invoice.withholding_amount / Decimal('0.05'))  # القيمة قبل الضريبة
                month_withholding_amount += float(invoice.withholding_amount)
        
        monthly_data.append({
            'month': month,
            'month_name': month_start.strftime('%B'),
            'total_sales': month_total_sales,
            'vat_sales': month_vat_sales,
            'vat_amount': month_vat_amount,
            'withholding_sales': month_withholding_sales,
            'withholding_amount': month_withholding_amount
        })
    
    # الحصول على بيانات الشركة من الإعدادات
    company_name = SystemSettings.get_setting('company_name', 'اسم الشركة')
    tax_number = SystemSettings.get_setting('tax_number', '000000000')
    company_address = SystemSettings.get_setting('company_address', 'عنوان الشركة')
    
    # تصدير إذا طُلب
    export_format = request.args.get('export')
    if export_format == 'pdf' and PDF_AVAILABLE:
        try:
            # تحضير البيانات للـ PDF
            pdf_data = {
                'company_name': company_name,
                'tax_number': tax_number,
                'company_address': company_address,
                'total_sales': total_sales,
                'total_vat_sales': total_vat_sales,
                'total_vat_amount': total_vat_amount,
                'total_withholding_sales': total_withholding_sales,
                'total_withholding_amount': total_withholding_amount,
                'monthly_data': monthly_data,
                'current_year': current_year
            }
            
            # إنشاء PDF
            pdf_buffer = pdf_generator.generate_tax_declaration_pdf(pdf_data)
            
            # إرسال الملف
            filename = f"tax_declaration_{current_year}.pdf"
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        except Exception as e:
            flash(f'خطأ في إنشاء PDF: {str(e)}', 'error')
    elif export_format == 'pdf':
        flash('تصدير PDF غير متاح حالياً', 'warning')
    
    return render_template('reports/tax_declaration.html',
                         total_sales=total_sales,
                         total_vat_sales=total_vat_sales,
                         total_vat_amount=total_vat_amount,
                         total_withholding_sales=total_withholding_sales,
                         total_withholding_amount=total_withholding_amount,
                         monthly_data=monthly_data,
                         current_year=current_year,
                         current_date=datetime.utcnow(),
                         company_name=company_name,
                         tax_number=tax_number,
                         company_address=company_address)

@reports_bp.route('/reports/monthly-summary')
@login_required
@permission_required('view_reports')
def monthly_summary():
    """ملخص شهري"""
    # الحصول على الشهر والسنة من المعاملات
    month = int(request.args.get('month', datetime.utcnow().month))
    year = int(request.args.get('year', datetime.utcnow().year))
    
    # تحديد بداية ونهاية الشهر
    month_start = datetime(year, month, 1)
    if month == 12:
        month_end = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = datetime(year, month + 1, 1) - timedelta(days=1)
    
    # استعلام فواتير الشهر
    monthly_invoices = Invoice.query.filter(
        Invoice.invoice_date >= month_start,
        Invoice.invoice_date <= month_end,
        Invoice.is_cancelled == False
    ).all()
    
    # حساب الإحصائيات
    total_invoices = len(monthly_invoices)
    total_sales = sum(inv.subtotal for inv in monthly_invoices)
    total_vat = sum(inv.vat_amount for inv in monthly_invoices)
    total_withholding = sum(inv.withholding_amount for inv in monthly_invoices)
    total_revenue = sum(inv.total_amount for inv in monthly_invoices)
    
    # أفضل العملاء
    customer_stats = {}
    for invoice in monthly_invoices:
        if invoice.customer_name not in customer_stats:
            customer_stats[invoice.customer_name] = {
                'invoices_count': 0,
                'total_amount': 0
            }
        customer_stats[invoice.customer_name]['invoices_count'] += 1
        customer_stats[invoice.customer_name]['total_amount'] += invoice.total_amount
    
    top_customers = sorted(customer_stats.items(), 
                          key=lambda x: x[1]['total_amount'], 
                          reverse=True)[:10]
    
    return render_template('reports/monthly_summary.html',
                         month=month,
                         year=year,
                         month_name=month_start.strftime('%B'),
                         monthly_invoices=monthly_invoices,
                         total_invoices=total_invoices,
                         total_sales=total_sales,
                         total_vat=total_vat,
                         total_withholding=total_withholding,
                         total_revenue=total_revenue,
                         top_customers=top_customers)

@reports_bp.route('/reports/yearly-summary')
@login_required
@permission_required('view_reports')
def yearly_summary():
    """ملخص سنوي"""
    # الحصول على السنة من المعاملات
    year = int(request.args.get('year', datetime.utcnow().year))
    
    # تحديد بداية ونهاية السنة
    year_start = datetime(year, 1, 1)
    year_end = datetime(year, 12, 31)
    
    # استعلام فواتير السنة
    yearly_invoices = Invoice.query.filter(
        Invoice.invoice_date >= year_start,
        Invoice.invoice_date <= year_end,
        Invoice.is_cancelled == False
    ).all()
    
    # حساب الإحصائيات السنوية
    total_invoices = len(yearly_invoices)
    total_sales = sum(inv.subtotal for inv in yearly_invoices)
    total_vat = sum(inv.vat_amount for inv in yearly_invoices)
    total_withholding = sum(inv.withholding_amount for inv in yearly_invoices)
    total_revenue = sum(inv.total_amount for inv in yearly_invoices)
    
    # إحصائيات شهرية
    monthly_stats = []
    for month in range(1, 13):
        month_start = datetime(year, month, 1)
        if month == 12:
            month_end = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = datetime(year, month + 1, 1) - timedelta(days=1)
        
        month_invoices = [inv for inv in yearly_invoices 
                         if month_start.date() <= inv.invoice_date <= month_end.date()]
        
        monthly_stats.append({
            'month': month,
            'month_name': month_start.strftime('%B'),
            'invoices_count': len(month_invoices),
            'sales': sum(inv.subtotal for inv in month_invoices),
            'vat': sum(inv.vat_amount for inv in month_invoices),
            'withholding': sum(inv.withholding_amount for inv in month_invoices),
            'revenue': sum(inv.total_amount for inv in month_invoices)
        })
    
    # الحصول على بيانات الشركة من الإعدادات
    company_name = SystemSettings.get_setting('company_name', 'اسم الشركة')
    tax_number = SystemSettings.get_setting('tax_number', '000000000')
    company_address = SystemSettings.get_setting('company_address', 'عنوان الشركة')
    
    # ترتيب الشهور حسب المبيعات (أفضل الشهور)
    top_months = sorted([m for m in monthly_stats if m['sales'] > 0], 
                       key=lambda x: x['sales'], reverse=True)
    
    # تصدير إذا طُلب
    export_format = request.args.get('export')
    if export_format == 'pdf' and PDF_AVAILABLE:
        try:
            # تحضير البيانات للـ PDF
            pdf_data = {
                'company_name': company_name,
                'tax_number': tax_number,
                'company_address': company_address,
                'year': year,
                'year_stats': {
                    'total_invoices': total_invoices,
                    'total_sales': total_sales,
                    'total_vat': total_vat,
                    'total_withholding': total_withholding,
                    'total_taxes': total_vat + total_withholding,
                    'total_revenue': total_revenue
                },
                'monthly_stats': monthly_stats,
                'top_months': top_months
            }
            
            # إنشاء PDF
            pdf_buffer = pdf_generator.generate_yearly_summary_pdf(pdf_data)
            
            # إرسال الملف
            filename = f"yearly_summary_{year}.pdf"
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        except Exception as e:
            flash(f'خطأ في إنشاء PDF: {str(e)}', 'error')
    elif export_format == 'pdf':
        flash('تصدير PDF غير متاح حالياً', 'warning')
    
    return render_template('reports/yearly_summary.html',
                         year=year,
                         yearly_invoices=yearly_invoices,
                         total_invoices=total_invoices,
                         total_sales=total_sales,
                         total_vat=total_vat,
                         total_withholding=total_withholding,
                         total_revenue=total_revenue,
                         monthly_stats=monthly_stats,
                         top_months=top_months,
                         company_name=company_name,
                         tax_number=tax_number,
                         company_address=company_address)

@reports_bp.route('/api/reports/chart-data')
@login_required
@permission_required('view_reports')
def chart_data():
    """بيانات الرسوم البيانية"""
    period = request.args.get('period', 'month')  # month, quarter, year
    
    if period == 'month':
        # بيانات آخر 12 شهر
        months_data = []
        for i in range(12):
            date = datetime.now().replace(day=1) - timedelta(days=30*i)
            month_start = date.replace(day=1)
            if i == 0:
                month_end = datetime.now().date()
            else:
                next_month = month_start.replace(month=month_start.month + 1) if month_start.month < 12 else month_start.replace(year=month_start.year + 1, month=1)
                month_end = (next_month - timedelta(days=1)).date()
            
            invoices = Invoice.query.filter(
                Invoice.invoice_date >= month_start.date(),
                Invoice.invoice_date <= month_end,
                Invoice.is_cancelled == False
            ).all()
            
            month_data = calculate_report_totals(invoices)
            month_data['period'] = month_start.strftime('%Y-%m')
            months_data.append(month_data)
        
        return jsonify(list(reversed(months_data)))
    
    return jsonify([])

def calculate_report_totals(invoices):
    """حساب إجماليات التقرير"""
    total_sales = sum(inv.subtotal for inv in invoices)
    total_vat = sum(inv.vat_amount for inv in invoices)
    total_withholding = sum(inv.withholding_amount for inv in invoices)
    total_amount = sum(inv.total_amount for inv in invoices)
    
    # تفاصيل المنتجات
    product_details = {}
    for invoice in invoices:
        for item in invoice.items:
            product_name = item.product.name
            if product_name not in product_details:
                product_details[product_name] = {
                    'quantity': 0,
                    'amount': 0,
                    'tax_type': item.product.tax_type.value,
                    'tax_amount': 0
                }
            
            line_total = item.get_line_total()
            product_details[product_name]['quantity'] += float(item.quantity)
            product_details[product_name]['amount'] += float(line_total)
            product_details[product_name]['tax_amount'] += float(item.get_tax_amount())
    
    return {
        'total_invoices': len(invoices),
        'total_sales': float(total_sales),
        'total_vat': float(total_vat),
        'total_withholding': float(total_withholding),
        'total_amount': float(total_amount),
        'product_details': product_details,
        'vat_invoices': len([inv for inv in invoices if inv.vat_amount > 0]),
        'withholding_invoices': len([inv for inv in invoices if inv.withholding_amount > 0])
    }

def create_pdf_report(report_data):
    """إنشاء تقرير PDF"""
    if not REPORTLAB_AVAILABLE:
        return None
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # إعداد الخطوط العربية (إذا كانت متوفرة)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # وسط
    )
    
    # محتوى التقرير
    story = []
    
    # العنوان
    story.append(Paragraph("تقرير الإقرارات الضريبية", title_style))
    story.append(Spacer(1, 12))
    
    # معلومات التقرير
    info_data = [
        ['الفترة:', f"{report_data['period_start']} إلى {report_data['period_end']}"],
        ['تاريخ الإنشاء:', report_data['generated_at'].strftime('%Y-%m-%d %H:%M')],
        ['أنشأ بواسطة:', report_data['generated_by']]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 12))
    
    # الإجماليات
    totals_data = [
        ['البيان', 'المبلغ (جنيه)'],
        ['إجمالي المبيعات (قبل الضريبة)', f"{report_data['total_sales']:,.2f}"],
        ['ضريبة القيمة المضافة (14%)', f"{report_data['total_vat']:,.2f}"],
        ['ضريبة الخصم والإضافة (5%)', f"{report_data['total_withholding']:,.2f}"],
        ['إجمالي المبلغ مع الضرائب', f"{report_data['total_amount']:,.2f}"]
    ]
    
    totals_table = Table(totals_data, colWidths=[3*inch, 2*inch])
    totals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(totals_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def create_excel_report(report_data):
    """إنشاء تقرير Excel"""
    if not OPENPYXL_AVAILABLE:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير الضريبي"
    
    # تنسيق الخلايا
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=16)
    
    # العنوان
    ws['A1'] = 'تقرير الإقرارات الضريبية'
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    # معلومات التقرير
    row = 3
    ws[f'A{row}'] = 'الفترة:'
    ws[f'B{row}'] = f"{report_data['period_start']} إلى {report_data['period_end']}"
    
    row += 1
    ws[f'A{row}'] = 'تاريخ الإنشاء:'
    ws[f'B{row}'] = report_data['generated_at'].strftime('%Y-%m-%d %H:%M')
    
    row += 1
    ws[f'A{row}'] = 'أنشأ بواسطة:'
    ws[f'B{row}'] = report_data['generated_by']
    
    # الإجماليات
    row += 3
    ws[f'A{row}'] = 'البيان'
    ws[f'B{row}'] = 'المبلغ (جنيه)'
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'إجمالي المبيعات (قبل الضريبة)'
    ws[f'B{row}'] = report_data['total_sales']
    
    row += 1
    ws[f'A{row}'] = 'ضريبة القيمة المضافة (14%)'
    ws[f'B{row}'] = report_data['total_vat']
    
    row += 1
    ws[f'A{row}'] = 'ضريبة الخصم والإضافة (5%)'
    ws[f'B{row}'] = report_data['total_withholding']
    
    row += 1
    ws[f'A{row}'] = 'إجمالي المبلغ مع الضرائب'
    ws[f'B{row}'] = report_data['total_amount']
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    
    # تفاصيل المنتجات
    if report_data['product_details']:
        row += 3
        ws[f'A{row}'] = 'تفاصيل المنتجات'
        ws[f'A{row}'].font = title_font
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ['المنتج', 'الكمية', 'المبلغ', 'الضريبة']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
        
        for product_name, details in report_data['product_details'].items():
            row += 1
            ws[f'A{row}'] = product_name
            ws[f'B{row}'] = details['quantity']
            ws[f'C{row}'] = details['amount']
            ws[f'D{row}'] = details['tax_amount']
    
    # تنسيق الأعمدة
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_report_pdf(report_data, report_id):
    """تصدير التقرير كـ PDF وحفظه"""
    if not REPORTLAB_AVAILABLE:
        return None
    
    # إنشاء مجلد التقارير إذا لم يكن موجوداً
    reports_dir = os.path.join('instance', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    filename = f'tax_report_{report_id}.pdf'
    filepath = os.path.join(reports_dir, filename)
    
    pdf_buffer = create_pdf_report(report_data)
    if pdf_buffer:
        with open(filepath, 'wb') as f:
            f.write(pdf_buffer.getvalue())
        return filepath
    
    return None

def export_report_excel(report_data, report_id):
    """تصدير التقرير كـ Excel وحفظه"""
    if not OPENPYXL_AVAILABLE:
        return None
    
    # إنشاء مجلد التقارير إذا لم يكن موجوداً
    reports_dir = os.path.join('instance', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    filename = f'tax_report_{report_id}.xlsx'
    filepath = os.path.join(reports_dir, filename)
    
    excel_buffer = create_excel_report(report_data)
    if excel_buffer:
        with open(filepath, 'wb') as f:
            f.write(excel_buffer.getvalue())
        return filepath
    
    return None
